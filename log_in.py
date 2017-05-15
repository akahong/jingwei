#encoding:utf-8
import unittest
from appium import webdriver
import time
from ddt import ddt,data,unpack
import MySQLdb #链接数据库
import re #正则表达式
import hashlib #md5加密模块
import types #类型判断类
from openpyxl import Workbook,load_workbook

#@ddt
class JingWei_LogIn(unittest.TestCase):
    global lwb
    def setUp(self):
        desired_caps={}
        desired_caps['platformName']='Android'
        desired_caps['platformVesion']='6.0'
        desired_caps['deviceName']='PLK-TL01H'
        desired_caps['appPackage']='com.jingwei.card'
        desired_caps['appActivity']='com.jingwei.card.LogoActivity'
        desired_caps['unicodeKeyboard']='True'
        desired_caps['resetKeyboard']='True'

        self.wb=webdriver.Remote('http://localhost:4723/wd/hub',desired_caps)
        time.sleep(5)

        self.conn=MySQLdb.Connect(host='10.8.64.53',
                             port=3306,
                             user='work@RR',
                             passwd='Geeker4ZolZ',
                             db='card',
                             charset='utf8')
        self.cursor=self.conn.cursor()

    #md5加密密码
    def md5(self,str):
        if type(str) is types.StringType: #判断传过来的参数是否为字符串
            m=hashlib.md5()  # 创建md5对象
            m.update(str) # 对password加密
            return m.hexdigest() # 获取加密字符串
        else:
            return ''

    #执行sql语句
    def sql(self,sql):
        self.cursor.execute(sql) #执行查询语句
        self.conn.commit() #提交查询语句
        print self.cursor.fetchone()
        return self.cursor.rowcount #返回查询条数

    #判断是否登录成功
    def if_login(self, sheet, i):
        self.wb.find_element_by_id('loginButton').click()  # 点击‘登录’按钮
        try:
            if self.wb.find_element_by_id('loginButton').is_displayed(): #以页面是否存在登录按钮来判断是否登录成功
                sheet['F'+str(i+2)].value='登录失败'
                if sheet['E'+str(i+2)].value==sheet['F'+str(i+2)].value:
                    sheet['G'+str(i+2)].value='pass'
                else:
                    sheet['G'+str(i+2)].value='false'                  
        except Exception, e:
            self.wb.find_element_by_id('radio_me').click() #点击‘我’
            self.wb.find_element_by_id('idManager').click() #点击‘账号管理’
            self.wb.find_element_by_id('btn_logout').click() #点击‘退出登录’
            self.wb.find_element_by_id('rightButton').click() #点击toast中的‘确定’
            sheet['F'+str(i+2)].value='登录成功' #判断是否登录成功
            if sheet['E'+str(i+2)].value==sheet['F'+str(i+2)].value:
                sheet['G'+str(i+2)].value='pass'
            else:
                sheet['G'+str(i+2)].value='false'               
        
    #获取测试数据
    def get_data(self):
        self.lwb=load_workbook('login_test_data.xlsx')
        sheet=self.lwb.get_sheet_by_name('login_test_data')
        return sheet
            
    #清空密码输入栏
    #def clear(self,el):
        #el.click() #选中输入框
        #self.wb.sendKeyEvent(123) #将光标移到最后
        #context=el.get_attribute('text') #获取字符串长度
        #for i in range(0,len(context)):
            #self.wb.keyevent(67) #一个个的删除
            
            
    #密码不正确
    def password_if_correct(self,sheet,i):
        sheet['F'+str(i+2)].value='密码不正确'
        if sheet['E'+str(i+2)].value==sheet['F'+str(i+2)].value:
            sheet['G'+str(i+2)].value='pass'
        else:
            sheet['G'+str(i+2)].value='false' 
            
    #账号未注册
    def account_if_login(self,account,sheet,i):
        if '@' in account:
            sheet['F'+str(i+2)].value='邮箱未注册'
        else:
            sheet['F'+str(i+2)].value='手机号未注册'
        if sheet['E'+str(i+2)].value==sheet['F'+str(i+2)].value:
            sheet['G'+str(i+2)].value='pass'
        else:
            sheet['G'+str(i+2)].value='false'        
                   

    #登录页面-登录
    def test_denglu(self):
        sheet = self.get_data() #获取测试数据表
        for i in range(sheet.max_row-1): #获取sheet的行数，进行循环
            account=str(sheet['C'+str(i+2)].value)
            password=str(sheet['D'+str(i+2)].value)
            self.wb.find_element_by_id('email').send_keys(account) # 输入账号
            try:
                #判断输入的是否为手机号
                if re.match(r'1(3|5|7|8)\d\d{8}',account):
                    sql_mobile='select mobile from account_mobile where mobile=%s'%account
                    if self.sql(sql_mobile)==1:  #判断手机是否注册
                        el=self.wb.find_element_by_id('passwordET') #定位到密码输入框      
                        el.send_keys(password) # 输入密码
                        #el.reset()
                        #self.clear(el) #调用清空密码方法
                        psw=self.md5(password) #调用加密方法对密码进行加密
                        sql_passwd_mobile="select password from account_mobile where mobile = '%s' and password='%s'"%(account, psw)
                        if self.sql(sql_passwd_mobile)==1:
                            self.if_login(sheet, i) #调用判断是否登录成功的方法
                        else:
                            self.password_if_correct(sheet,i) #调用判密码不正确的方法 
                            self.wb.find_element_by_id('clearPwd').click() #清空密码输入框
                    else:
                        self.account_if_login(account,sheet,i)  #账号未注册
                        
    
                #判断输入是否为邮箱
                elif re.match(r'\d+@\w+.com',account):
                    sql_email="select email from account_email where email='%s'" %account
                    if self.sql(sql_email)== 1: #判断邮箱是否注册
                        self.wb.find_element_by_id('passwordET').send_keys(password)  # 输入密码
                        psw_email=self.md5(password) #调用加密方法对密码进行加密
                        print 'psw:'+psw_email
                        sql_passwd_email= "select password from account_email where email='%s' and password='%s'" %(account,psw_email)
                        if self.sql(sql_passwd_email) == 1: #判断对应的密码是否正确
                            self.if_login(sheet, i) #调用判断是否登录成功的方法                          
                        else:
                            self.password_if_correct(sheet,i) #调用判断密码是否正确的方法 
                            self.wb.find_element_by_id('clearPwd').click() #清空密码输入框
                    else:
                        self.account_if_login(account,sheet,i) #账号未注册                      
                else:
                    sheet['F'+str(i+2)].value='账号格式不正确'
                    if sheet['E'+str(i+2)].value==sheet['F'+str(i+2)].value:
                        sheet['G'+str(i+2)].value='pass'
                    else:
                        sheet['G'+str(i+2)].value='false'                    
            except Exception,e:
                print e
        self.lwb.save('result.xlsx')
    
    #登录页面-先试试
    #def test_xianshishi(self):
        #self.wb.find_element_by_id('rightTextView').click()
        #self.wb.find_element_by_id('leftButton').click()
        #try:
            #if self.wb.find_element_by_id('radio_home').is_displayed(): #判断页面是否显示‘名片夹’按钮
                #print '先试试页面跳转成功'
        #except Exception,e:
            #print '先试试页面跳转失败'

    
                
    #清除方法
    def tearDown(self):
        self.wb.quit() #关闭appium链接
        self.cursor.close() #关闭游标
        self.conn.close() #关闭数据库链接

if __name__ == '__main__':
    unittest.main()


#adb shell pm list packages -3 显示app应用的package名